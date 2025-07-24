# -*- coding: utf-8 -*-
from __future__ import unicode_literals

import logging
from datetime import datetime
from decimal import Decimal, InvalidOperation
from io import BytesIO
from string import Formatter
from typing import List, Dict

from lxml import etree
from openpyxl import Workbook
from openpyxl.cell import WriteOnlyCell
from openpyxl.styles import Font
from openpyxl.styles.alignment import Alignment
from openpyxl.styles.fills import PatternFill
from openpyxl.styles.named_styles import NamedStyle
from openpyxl.utils import get_column_letter, column_index_from_string
from openpyxl.worksheet.worksheet import Worksheet
from six import text_type

logger = logging.getLogger(__name__)


class CellRef:
    """
    Handy class to store cell reference and add sheet name when cast to Unicode, if needed.

    Returns references in ``sheet!column-row`` Excel style.
    """

    def __init__(self, target, row, col, sheet_title=None):
        self._target = target
        self.sheet_title: str = sheet_title or target._current_ws.title
        self.row = row
        self.col = col

    def __repr__(self):
        return self.__unicode__()

    def __unicode__(self):
        if self.sheet_title == self._target._current_ws.title:
            return u'{col}{row}'.format(
                sheet=self.sheet_title,
                col=get_column_letter(self.col + 1),
                row=self.row + 1
            )
        else:
            return u"'{sheet}'!{col}{row}".format(
                sheet=self.sheet_title,
                col=get_column_letter(self.col + 1),
                row=self.row + 1
            )


class XML2XLSXTarget:

    def __init__(self, write_only: bool = False, cell_names: List[str] | None = None):
        self.write_only: bool = write_only

        self.wb: Workbook = Workbook(write_only=write_only)
        if self.wb.sheetnames:
            std = self.wb['Sheet']
            self.wb.remove(std)
        self._current_ws: Worksheet | None = None
        self._row_buf: List[WriteOnlyCell] = []
        self._cell = None
        self._cell_type: str | None = None
        self._cell_date_format = None
        self._row: int = 0
        self._col: int = 0

        if cell_names and (
                "row" in cell_names or "sheet" in cell_names or "columns" in cell_names or
                "style" in
                cell_names):
            raise ValueError(u"Cell names 'row', 'columns', 'style' and 'sheet' are reserved")
        self.cell_names: List[str] = cell_names or []
        if "cell" not in self.cell_names:
            self.cell_names.append("cell")
        self._refs: Dict[str, CellRef | int] = {
            'row': 1,
            'col': 1,
        }

    @staticmethod
    def _parse_descriptor(descriptor: str) -> Dict[str, str | bool | int | float]:
        params: Dict[str, str] = dict([v.split(':') for v in descriptor.split(';') if v.strip()])
        return {param.strip(): XML2XLSXTarget.parse_type(value.strip()) for param, value in
                params.items()}

    @staticmethod
    def parse_type(value: str) -> bool | int | float | str | None:
        if value is None:
            return None
        elif value.lower() in ['true', 'false']:
            return bool(value.capitalize())
        else:
            try:
                return int(value)
            except ValueError:
                try:
                    return float(value)
                except ValueError:
                    return value

    @staticmethod
    def _get_font(desc) -> Font or None:
        return Font(**XML2XLSXTarget._parse_descriptor(desc))

    @staticmethod
    def _get_alignment(desc) -> Alignment or None:
        return Alignment(**XML2XLSXTarget._parse_descriptor(desc))

    @staticmethod
    def _get_fill(desc) -> PatternFill or None:
        params = XML2XLSXTarget._parse_descriptor(desc)
        if params['fill_type'] == 'solid':
            return PatternFill(**params)
        elif params['fill_type'] == 'gradient':
            raise NotImplementedError('Gradient fills are not supported')
            # return GradientFill(**params)
        return None

    def start(self, tag, attrib) -> None:

        if not self._current_ws:
            self._current_ws = self.wb.active

        if tag == 'sheet':
            index = int(attrib.get('index')) if 'index' in attrib else None
            self._current_ws = self.wb.create_sheet(
                title=attrib.get('title', None), index=index
            )
            self._row = 0
        elif tag == 'columns':
            start = column_index_from_string(attrib['start'])
            end = column_index_from_string(attrib.get('end', attrib['start']))
            for i in range(start, end + 1):
                self._current_ws.column_dimensions[
                    get_column_letter(i)
                ].width = int(attrib.get('width')) / 7.0
        elif tag == 'row':
            self._row_buf = []
            self._col = 0
        elif tag in self.cell_names:
            self._cell = WriteOnlyCell(self._current_ws)
            for attr, value in attrib.items():
                if attr == 'font':
                    self._cell.font = self._get_font(value)
                elif attr == 'fill':
                    self._cell.fill = self._get_fill(value)
                elif attr == 'alignment':
                    self._cell.alignment = self._get_alignment(value)
                elif attr == 'ref-id':
                    self._refs[value] = CellRef(self, self._row, self._col)
                elif attr == 'ref-append':
                    self._refs[value] = self._refs.get(value, [])
                    self._refs[value].append(CellRef(self, self._row, self._col))
                elif attr == 'fmt':
                    self._cell.number_format = value
                elif attr == 'rowspan':
                    self._current_ws.merge_cells(
                        start_row=self._row + 1, start_column=self._col + 1,
                        end_row=self._row + int(value), end_column=self._col + 1
                    )
                elif attr == 'colspan':
                    self._current_ws.merge_cells(
                        start_row=self._row + 1, start_column=self._col + 1,
                        end_row=self._row + 1, end_column=self._col + int(value)
                    )

            ctype = attrib.get('type', 'unicode')
            if ctype not in ['unicode', 'number', 'date']:
                raise ValueError(u'Unknown cell type {ctype}.'.format(
                    ctype=ctype,
                ))
            self._cell_type = ctype
            try:
                self._cell_date_format = attrib.get('date-fmt')
            except KeyError:
                raise ValueError(u"Specify 'date-fmt' attribute for 'date'"
                                 u" type")

        elif tag == 'style':
            style = NamedStyle(name=attrib['name'])
            if 'font' in attrib:
                style.font = self._get_font(attrib['font'])
            if 'fill' in attrib:
                style.fill = self._get_fill(attrib['fill'])
            self.wb.add_named_style(style)

    def data(self, data):
        if self._cell:
            if self._cell.value:
                # TODO: Szybki fix na to, że znakiunicode powodują przerwanie
                #  czytania data i rozbijają to na 2
                self._cell.value += data
            else:
                self._cell.value = data

    def end(self, tag):
        if tag == 'sheet':
            pass
        elif tag == 'row':
            self._current_ws.append(self._row_buf)
            self._row_buf = []
            self._row += 1
            self._refs['row'] = self._row + 1
        elif tag == 'cell':
            if self._cell_type == 'unicode' and isinstance(self._cell.value,
                                                           str) and self._cell.value:
                keys = [
                    e[1] for e in Formatter().parse(self._cell.value)
                    if e[1] is not None
                ]

                stringified = {
                    k: ', '.join(text_type(e) for e in self._refs.get(k, ''))
                    if hasattr(self._refs.get(k, ''), '__iter__')
                    else text_type(self._refs.get(k, ''))
                    for k in keys or []
                }
                self._cell.value = self._cell.value.format(**stringified)
            elif self._cell_type == 'number' and self._cell.value:
                try:
                    self._cell.value = Decimal(self._cell.value)
                except InvalidOperation:
                    pass
            elif self._cell_type == 'date' and self._cell.value:
                try:
                    self._cell.value = datetime.strptime(str(self._cell.value),
                                                         self._cell_date_format).date()
                except TypeError:
                    pass
            self._row_buf.append(self._cell)
            self._cell = None
            self._col += 1
            self._refs['col'] = self._col + 1

    def close(self):
        with BytesIO() as f:
            self.wb.save(f)
            f.seek(0)
            return f.getvalue()


def xml2xlsx(xml):
    """
    Converts XML in a proper format to a xlsx (MS Excel) file.

    The XML argument is **not** an Excel file in XML format.
    :param xml: A string with proper XML.
    :type xml: Unicode
    :return: Parsed XML that can be saved to a stream.
    """
    parser = etree.XMLParser(target=XML2XLSXTarget(), encoding='UTF-8',
                             remove_blank_text=True, huge_tree=True)
    return etree.XML(xml, parser, )
