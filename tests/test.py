# -*- coding: utf-8 -*-
from __future__ import unicode_literals

import io
import logging
import tempfile
import pytest
from datetime import date
from timeit import default_timer as timer


from openpyxl.reader.excel import load_workbook
from openpyxl.styles.alignment import Alignment
from openpyxl.styles.fills import PatternFill
from six import text_type
from ..xml2xlsx import xml2xlsx, XML2XLSXTarget, CellRef

logger = logging.getLogger(__name__)


class TestCellRef:

    @pytest.fixture(autouse=True)
    def setup(self):
        self.target = XML2XLSXTarget()

    def test_unicode_same_worksheet(self):
        self.target.start(tag='sheet', attrib={'title': 'test1'})
        cell = CellRef(self.target, 0, 0)
        assert text_type(cell) == u'A1'

    def test_unicode_far_column(self):
        self.target.start(tag='sheet', attrib={'title': 'test1'})
        cell = CellRef(self.target, 0, 26)
        assert text_type(cell) == u'AA1'

    def test_unicode_different_worksheet(self):
        self.target.start(tag='sheet', attrib={'title': 'test1'})
        cell = CellRef(self.target, 0, 0)
        self.target.end(tag='sheet')
        self.target.start(tag='sheet', attrib={'title': 'test2'})
        assert text_type(cell) == "'test1'!A1"


class TestXML2XLSX:
    def test_single_row(self):
        template = """
        <sheet title="test">
            <row>
                <cell>test cell</cell>
                <cell>test cell2</cell>
            </row>
        </sheet>

        """
        sheet = io.BytesIO(xml2xlsx(template))
        wb = load_workbook(sheet,read_only=True, data_only=True)
        assert len(wb.worksheets) == 1, "Created workbook should have only one sheet"
        assert "test" in wb.sheetnames, "Worksheet 'test' missing"
        ws = wb["test"]
        assert ws["A1"].value == "test cell"
        assert ws["B1"].value == "test cell2"

    def test_save_to_file(self):
        template = '<sheet title="test"></sheet>'
        f = tempfile.TemporaryFile()
        f.write(xml2xlsx(template))
        f.seek(0)
        wb = load_workbook(f)
        assert len(wb.worksheets) == 1, "Created workbook should have only one sheet"
        assert "test" in wb.sheetnames, "Worksheet 'test' missing"
        f.close()

    def test_xml_special_chars(self):
        template = """
        <sheet title="test">
            <row>
                <cell>2&lt;=3</cell>
            </row>
        </sheet>

        """
        sheet = io.BytesIO(xml2xlsx(template))
        wb = load_workbook(sheet)
        ws = wb["test"]
        assert ws["A1"].value == "2<=3"

    def test_cell_font(self):
        template = """
        <sheet title="test">
            <row>
                <cell font="size: 10; bold: True;">test cell</cell>
            </row>
        </sheet>

        """
        sheet = io.BytesIO(xml2xlsx(template))
        wb = load_workbook(sheet)

        assert len(wb.worksheets) == 1, "Created workbook should have only one sheet"
        assert "test" in wb.sheetnames, "Worksheet 'test' missing"
        ws = wb["test"]
        assert ws["A1"].font.size == 10, "Font size not set properly"
        assert ws["A1"].font.bold, "Font is not bold"

    def test_cell_fill(self):
        template = """
        <sheet title="test">
            <row>
                <cell fill="fill_type: solid; bgColor: 00BFBFBF">test</cell>
            </row>
        </sheet>
        """
        sheet = io.BytesIO(xml2xlsx(template))
        wb = load_workbook(sheet)
        assert "test" in wb.sheetnames, "Worksheet 'test' missing"
        ws = wb["test"]
        assert ws["A1"].fill.fill_type == 'solid'
        assert ws["A1"].fill.bgColor.rgb == "00BFBFBF"

    def test_unicode(self):
        template = """
        <sheet title="test">
            <row><cell>aąwźćńół</cell></row>
        </sheet>
        """
        sheet = io.BytesIO(xml2xlsx(template))
        wb = load_workbook(sheet)
        ws = wb["test"]
        assert ws["A1"].value == "aąwźćńół"

    def test_multiple_rows(self):
        template = """
        <sheet title="test">
            <row>
                <cell>test cell</cell>
            </row>
            <row>
                <cell>test cell2</cell>
            </row>
        </sheet>
        """
        sheet = io.BytesIO(xml2xlsx(template))
        wb = load_workbook(sheet)
        ws = wb["test"]
        assert ws["A1"].value == "test cell"
        assert ws["A2"].value == "test cell2"

    def test_cell_type_number(self):
        template = """
        <sheet title="test"><row><cell type="number">1123.4</cell></row>
        </sheet>
        """
        sheet = io.BytesIO(xml2xlsx(template))
        wb = load_workbook(sheet)
        ws = wb["test"]
        assert ws["A1"].value == 1123.4

    def test_cell_type_date(self):
        template = """
        <sheet title="test">
            <row><cell type="date" date-fmt="%d.%m.%Y">24.01.1981</cell></row>
        </sheet>
        """
        sheet = io.BytesIO(xml2xlsx(template))
        wb = load_workbook(sheet)
        ws = wb["test"]
        assert ws["A1"].value.date() == date(1981, 1, 24)

    def test_empty_cell_type_date(self):
        template = """
        <sheet title="test">
            <row><cell type="date" date-fmt="%d.%m.%Y"></cell></row>
        </sheet>
        """
        sheet = io.BytesIO(xml2xlsx(template))
        wb = load_workbook(sheet)
        ws = wb['test']
        assert ws['A1'].value is None

    def test_cell_number_format(self):
        template = """
        <sheet title="test">
            <row>
                <cell type="number" fmt="# ##0.000;[RED]# ##0.000">
                   1
                </cell>
            </row>
        </sheet>
        """
        sheet = io.BytesIO(xml2xlsx(template))
        wb = load_workbook(sheet)
        ws = wb["test"]
        assert ws["A1"].number_format == '# ##0.000;[RED]# ##0.000'

    def test_cell_alignment(self):
        template = """
        <sheet title="test">
            <row>
                <cell alignment="horizontal: general">
                   1
                </cell>
            </row>
        </sheet>
        """
        sheet = io.BytesIO(xml2xlsx(template))
        wb = load_workbook(sheet)
        ws = wb["test"]
        assert ws["A1"].alignment.horizontal == 'general'

    def test_cell_ref_id(self):
        template = """
        <sheet title="test">
            <row><cell ref-id="refcell">XXXX</cell></row>
            <row><cell>{refcell}</cell></row>
        </sheet>
        """
        sheet = io.BytesIO(xml2xlsx(template))
        wb = load_workbook(sheet)
        ws = wb["test"]
        assert ws["A2"].value == "A1"

    def test_cell_ref_id_inexistent(self):
        template = """
        <sheet title="test">
            <row><cell>{refcell}</cell></row>
        </sheet>
        """
        sheet = io.BytesIO(xml2xlsx(template))
        wb = load_workbook(sheet)
        ws = wb["test"]
        assert ws["A1"].value is None

    def test_cell_ref_id_different_worksheet(self):
        template = """
        <workbook>
            <sheet title="test">
                <row><cell ref-id="refcell">XXXX</cell></row>
                <row><cell>{refcell}</cell></row>
            </sheet>
            <sheet title="test2">
                <row><cell>{refcell}</cell></row>
            </sheet>
        </workbook>
        """
        sheet = io.BytesIO(xml2xlsx(template))
        wb = load_workbook(sheet)
        assert wb['test2']["A1"].value == "'test'!A1"

    def test_cell_ref_col(self):
        template = """
        <sheet title="test">
            <row><cell>{col}</cell><cell>{col}</cell></row>
        </sheet>
        """
        sheet = io.BytesIO(xml2xlsx(template))
        wb = load_workbook(sheet)
        ws = wb["test"]
        assert ws["A1"].value == "1"
        assert ws["B1"].value == "2"

    def test_cell_ref_row(self):
        template = """
        <sheet title="test">
            <row><cell>{row}</cell></row>
            <row><cell>{row}</cell></row>
        </sheet>
        """
        sheet = io.BytesIO(xml2xlsx(template))
        wb = load_workbook(sheet)
        ws = wb["test"]
        assert ws["A1"].value == "1"
        assert ws["A2"].value == "2"

    def test_cell_ref_append(self):
        template = """
        <sheet title="test">
            <row><cell ref-append="my-list">ABC</cell></row>
            <row><cell ref-append="my-list">DEFG</cell></row>
            <row><cell>{my-list}</cell></row>
        </sheet>
        """
        sheet = io.BytesIO(xml2xlsx(template))
        wb = load_workbook(sheet)
        ws = wb["test"]
        assert ws["A3"].value == "A1, A2"

    def test_sheet_index_attrib(self):
        template = """
        <workbook>
            <sheet title="test">
            </sheet>
            <sheet title="test2" index="0">
            </sheet>
        </workbook>
        """
        sheet = io.BytesIO(xml2xlsx(template))
        wb = load_workbook(sheet)
        assert wb.sheetnames == ["test2", "test"]

    def test_column_width(self):
        template = """
        <workbook>
            <sheet title="test">
                <columns start="A" end="D" width="14"/>
            </sheet>
        </workbook>
        """
        sheet = io.BytesIO(xml2xlsx(template))
        wb = load_workbook(sheet)
        ws = wb["test"]
        for col in ['A', 'B', 'C', 'D']:
            assert ws.column_dimensions[col].width == 2
        assert ws.column_dimensions['E'].width is not 14

    def test_named_style(self):
        template = """
        <workbook>
            <style name="test"/>
            <sheet title="test"/>
        </workbook>
        """
        sheet = io.BytesIO(xml2xlsx(template))
        wb = load_workbook(sheet)
        assert 'test' in wb.style_names

    def test_named_style_with_cell(self):
        template = """
        <workbook>
            <style name="test"/>
            <sheet title="test"><row><cell>a</cell></row></sheet>
        </workbook>
        """
        sheet = io.BytesIO(xml2xlsx(template))
        wb = load_workbook(sheet)
        assert 'test' in wb.style_names

    def test_named_style_font(self):
        template = """
        <workbook>
            <style name="test" font="bold: True;"/>
            <sheet title="test"/>
        </workbook>
        """
        sheet = io.BytesIO(xml2xlsx(template))
        wb = load_workbook(sheet)
        assert 'test' in wb.style_names
        style = wb._named_styles['test']
        assert style.font.bold

    def test_named_style_fill_solid(self):
        template = """
        <workbook>
            <style name="test" fill="fill_type: solid; fgColor: BFBFBF"/>
            <sheet title="test"/>
        </workbook>
        """
        sheet = io.BytesIO(xml2xlsx(template))
        wb = load_workbook(sheet)
        style = wb._named_styles['test']
        assert isinstance(style.fill, PatternFill)
        assert style.fill.fill_type == 'solid'
        assert style.fill.fgColor.rgb == '00BFBFBF'

    def test__parse_descriptor_bool(self):
        descriptor = "test: True"
        params = XML2XLSXTarget._parse_descriptor(descriptor)
        assert params == {'test': True}

    def test__parse_descriptor_int(self):
        descriptor = "test: 123"
        params = XML2XLSXTarget._parse_descriptor(descriptor)
        assert params == {'test': 123}

    def test__parse_descriptor_float(self):
        descriptor = "test: 123.3"
        params = XML2XLSXTarget._parse_descriptor(descriptor)
        assert params == {'test': 123.3}

    def test__parse_descriptor_string(self):
        descriptor = "test:  abc"
        params = XML2XLSXTarget._parse_descriptor(descriptor)
        assert params == {'test': 'abc'}

    def test__parse_descriptor_multiple(self):
        descriptor = "test: True; test2: 1; test3: 3.0; test4: abc;"
        params = XML2XLSXTarget._parse_descriptor(descriptor)
        assert params == {
            'test': True, 'test2': 1, 'test3': 3.0, 'test4': 'abc'
        }

    def test__get_font(self):
        descriptor = "size: 10"
        font = XML2XLSXTarget._get_font(descriptor)
        assert font.size == 10

    def test__get_alignment(self):
        descriptor = "horizontal: general"
        alignment = XML2XLSXTarget._get_alignment(descriptor)
        assert isinstance(alignment, Alignment)
        assert alignment.horizontal == 'general'

    def test__get_fill_solid(self):
        descriptor = "fill_type: solid"
        fill = XML2XLSXTarget._get_fill(descriptor)
        assert isinstance(fill, PatternFill)
        assert fill.patternType == 'solid'


@pytest.mark.performance
class TestXML2XLSXPerformance:

    def test_single_sheet(self):
        inhalt = [
            '\n<row>' + '<cell>test</cell>' * 100 + '</row>'
            for _ in range(1000)
        ]
        template = u'<workbook><sheet title="test">%s</sheet></workbook>' % (
            ''.join(inhalt)
        )
        start = timer()
        io.BytesIO(xml2xlsx(template))
        end = timer()
        logger.info('Single sheet performace test result: %s', end-start)


if __name__ == '__main__':
    pytest.main()
